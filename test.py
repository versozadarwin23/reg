import re
from openpyxl import Workbook, load_workbook
import os
import requests
from bs4 import BeautifulSoup
import time
import random
from concurrent.futures import ThreadPoolExecutor
import threading

os.system("clear")

SUCCESS = "✅"
FAILURE = "❌"
INFO = "ℹ️"
WARNING = "⚠️"
LOADING = "⏳"
custom_password_base = None
lock = threading.Lock()

def load_names_from_file(file_path):
    with open(file_path, 'r') as file:
        return [line.strip() for line in file.readlines()]

def get_names(account_type, gender):
    if account_type == 1:
        male_first_names = load_names_from_file("first_name.txt")
        last_names = load_names_from_file("last_name.txt")
        female_first_names = []
    else:
        male_first_names = []
        female_first_names = load_names_from_file('path_to_female_first_names.txt')
        last_names = load_names_from_file('path_to_last_names.txt')

    firstname = random.choice(male_first_names if gender == 1 else female_first_names)
    lastname = random.choice(last_names)
    return firstname, lastname

def generate_random_phone_number():
    random_number = str(random.randint(1000000, 9999999))
    third = random.randint(0, 4)
    forth = random.randint(1, 7)
    return f"9{third}{forth}{random_number}"

def generate_random_password():
    global custom_password_base
    # Never ask for input here
    six_digit = str(random.randint(100000, 999999))
    return custom_password_base + six_digit


def generate_user_details(account_type, gender, password=None):
    firstname, lastname = get_names(account_type, gender)
    year = random.randint(1978, 2001)
    date = random.randint(1, 28)
    month = random.randint(1, 12)
    if password is None:
        password = generate_random_password()
    phone_number = generate_random_phone_number()
    return firstname, lastname, date, year, month, phone_number, password

def clean_excel_text(text):
    if not isinstance(text, str):
        text = str(text)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

def save_to_xlsx(filename, data):
    with lock:
        try:
            cleaned_data = [clean_excel_text(item) for item in data]

            if os.path.isfile(filename):
                workbook = load_workbook(filename)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['NAME', 'USERNAME', 'PASSWORD', 'ACCOUNT LINK'])

            sheet.append(cleaned_data)
            workbook.save(filename)
        except Exception as e:
            print(f"{FAILURE} Error saving to {filename}: {e}. Retrying...")

def create_fbunconfirmed(account_type, usern, gender, password=None, email=None, retry_if_checkpoint=True):
    global custom_password_base
    firstname, lastname, date, year, month, phone_number, used_password = generate_user_details(account_type, gender, password)

    url = "https://m.facebook.com/reg"
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://m.facebook.com/reg",
        "Connection": "keep-alive",
        "X-FB-Connection-Type": "MOBILE.LTE",
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        'x-fb-connection-type': 'Unknown',
        'accept-encoding': 'gzip, deflate',
        'content-type': 'application/x-www-form-urlencoded',
        'x-fb-http-engine': 'Liger',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; CPH1903 Build/O11019; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/70.0.3538.110 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/444.0.0.0.110;]',
    }

    session = requests.Session()

    try:
        response = session.get(url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")
        form = soup.find("form")
        if not form:
            with lock:
                print(f"{FAILURE} Failed to load registration form.")
            return

        action_url = requests.compat.urljoin(url, form["action"]) if form.has_attr("action") else url
        inputs = form.find_all("input")

        email_or_phone = email
        data = {
            "firstname": firstname,
            "lastname": lastname,
            "birthday_day": str(date),
            "birthday_month": str(month),
            "birthday_year": str(year),
            "reg_email__": email_or_phone,
            "sex": str(gender),
            "encpass": used_password,
            "submit": "Sign Up"
        }

        for inp in inputs:
            if inp.has_attr("name") and inp["name"] not in data:
                data[inp["name"]] = inp["value"] if inp.has_attr("value") else ""

        post_response = session.post(action_url, headers=headers, data=data)

        refresh_url = "https://m.facebook.com/reg"
        refreshed_response = session.get(refresh_url, headers=headers)
        refreshed_soup = BeautifulSoup(refreshed_response.text, "html.parser")

        # Check if checkpointed
        checkpoint_form = refreshed_soup.find('form', action=lambda x: x and 'checkpoint' in x)
        if checkpoint_form or "checkpoint" in post_response.url.lower():
            if retry_if_checkpoint:
                new_password = generate_random_password()
                return create_fbunconfirmed(account_type, usern, gender, password=new_password, email=email, retry_if_checkpoint=False)
            else:
                print(f"{FAILURE} {email_or_phone} failed twice. Skipping.")
                return

        if "c_user" in session.cookies:
            uid = session.cookies.get("c_user")
            profile_id = f"https://www.facebook.com/profile.php?id={uid}"
            full_name = f"{firstname} {lastname}"
            filename = "/storage/emulated/0/Acc_Created.xlsx"
            data_to_save = [full_name, email_or_phone, used_password, profile_id]
            save_to_xlsx(filename, data_to_save)
            with lock:
                info = f"\033[92m{SUCCESS}| {email_or_phone} | {used_password} |\033[0m"
                msg = f"{full_name}	{email_or_phone}	{used_password}	{profile_id}"
                print(info)
                try:
                    with open("/storage/emulated/0/fb_created_log.txt", "a", encoding="utf-8") as log_file:
                        log_file.write(msg + "\n")
                except Exception as e:
                    print(f"{FAILURE} Failed to write log: {e}")
        else:
            with lock:
                print(f"{FAILURE} {email_or_phone} creation failed. Account got blocked.")
    except Exception as e:
        print(f"{FAILURE} Error during creation: {e}")


def threaded_worker(index, account_type, gender, email):
    time.sleep(3 * index)  # delay start
    usern = f"user{index + 1}"
    os.system("clear")
    time.sleep(3)
    create_fbunconfirmed(account_type, usern, gender, email=email)

def main_with_threads():
    os.system("clear")
    try:
        max_create = int(input("🔢 Enter number of accounts to create: "))
        max_workers = int(input("🧵 Enter max workers (threads): "))
    except ValueError:
        print(f"{FAILURE} Invalid input.")
        return

    # ✅ Prompt password base once
    global custom_password_base
    custom_password_base = input("\033[92m🔐 Type your password: \033[0m").strip() or "promises"

    emails = []
    for i in range(max_create):
        email = input(f"📧 Enter email for account #{i + 1}: ")
        emails.append(email.strip())

    account_type = 1
    gender = 1

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for i in range(max_create):
            futures.append(executor.submit(threaded_worker, i, account_type, gender, emails[i]))

        for future in futures:
            future.result()

if __name__ == "__main__":
    main_with_threads()
