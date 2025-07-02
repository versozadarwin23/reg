import string
from openpyxl import Workbook, load_workbook
import os
import requests
import time
import random
import json
import concurrent.futures
import threading
import re # Import re for regular expressions
from urllib.parse import urljoin # Import urljoin for correctly handling relative URLs

xlsx_lock = threading.Lock()
console_lock = threading.Lock()
os.system("clear")

def random_device_model():
    models = [
        "Samsung-SM-S918B",
        "Xiaomi-2210132G",
        "OnePlus-CPH2451",
        "OPPO-CPH2207",
        "vivo-V2203",
        "realme-RMX3085",
        "Samsung-Galaxy-A54",
        "Samsung-SM-A146P",
        "Samsung-Galaxy-S23Ultra",
        "Samsung-SM-F946B",
        "Samsung-Galaxy-M34",
        "Xiaomi-23049PCD8G",
        "Xiaomi-Redmi-Note-12",
        "Xiaomi-POCO-X5Pro",
        "Xiaomi-2312DRA50G",
        "OnePlus-CPH2513",
        "OnePlus-CPH2581",
        "OnePlus-CPH2459",
        "OPPO-CPH2339",
        "OPPO-CPH2419",
        "OPPO-CPH2521",
        "vivo-V2140",
        "vivo-V2254",
        "vivo-V2230",
        "vivo-V2313A",
        "realme-RMX3612",
        "realme-RMX3571",
        "realme-RMX3761",
        "realme-RMX3491",
        "Huawei-ANE-LX2",
        "Huawei-JNY-LX1",
        "Huawei-ELS-NX9",
        "Huawei-CDY-NX9B",
        "Motorola-Moto-G73",
        "Motorola-XT2345-4",
        "Motorola-XT2303-2",
        "Infinix-X6815B",
        "Infinix-X6711",
        "Infinix-X676C",
        "TECNO-CK7n",
        "TECNO-CH9n",
        "TECNO-BD4h",
        "HONOR-ANY-AN00",
        "HONOR-MGA-AN00",
        "HONOR-LRA-AN00",
        "Lenovo-L78051",
        "Lenovo-K13-Note",
        "Google-Pixel-7",
        "Google-Pixel-6a",
        "Google-Pixel-5"
    ]
    return random.choice(models)


def random_device_id():
    ids = [
        "0f47e6d2-bb61-4bfc-80db-123456789001",
        "1a2b3c4d-5e6f-7a8b-9c0d-234567890002",
        "2b3c4d5e-6f7a-8b9c-0d1e-345678900003",
        "3c4d5e6f-7a8b-9c0d-1e2f-456789000004",
        "4d5e6f7a-8b9c-0d1e-2f3a-567890000005",
        "5e6f7a8b-9c0d-1e2f-3a4b-678900000006",
        "6f7a8b9c-0d1e-2f3a-4b5c-789000000007",
        "7a8b9c0d-1e2f-3a4b-5c6d-890000000008",
        "8b9c0d1e-2f3a-4b5c-6d7e-900000000009",
        "9c0d1e2f-3a4b-5c6d-7e8f-000000000010",
        "aa1bb2cc-3dd4-5ee6-7ff8-111111111011",
        "bb2cc3dd-4ee5-6ff7-8009-222222222012",
        "cc3dd4ee-5ff6-7008-9110-333333333013",
        "dd4ee5ff-6007-8119-0221-444444444014",
        "ee5ff600-7118-9220-1332-555555555015",
        "ff600711-8229-0331-2443-666666666016",
        "00611722-9330-1442-3554-777777777017",
        "11722833-0441-2553-4665-888888888018",
        "22833944-1552-3664-5776-999999999019",
        "33944a55-2663-4775-6887-000000000020",
        "44a55b66-3774-5886-7998-111111111021",
        "55b66c77-4885-6997-8009-222222222022",
        "66c77d88-5996-7008-9110-333333333023",
        "77d88e99-6007-8119-0221-444444444024",
        "88e990aa-7118-9220-1332-555555555025",
        "990aa1bb-8229-0331-2443-666666666026",
        "0aa1bb2c-9330-1442-3554-777777777027",
        "1bb2cc3d-0441-2553-4665-888888888028",
        "2cc3dd4e-1552-3664-5776-999999999029",
        "3dd4ee5f-2663-4775-6887-000000000030",
        "4ee5ff60-3774-5886-7998-111111111031",
        "5ff60071-4885-6997-8009-222222222032",
        "60071182-5996-7008-9110-333333333033",
        "71182293-6007-8119-0221-444444444034",
        "82293304-7118-9220-1332-555555555035",
        "93304415-8229-0331-2443-666666666036",
        "04415526-9330-1442-3554-777777777037",
        "15526637-0441-2553-4665-888888888038",
        "26637748-1552-3664-5776-999999999039",
        "37748859-2663-4775-6887-000000000040",
        "48859960-3774-5886-7998-111111111041",
        "59960071-4885-6997-8009-222222222042",
        "60071182-5996-7008-9110-333333333043",
        "71182293-6007-8119-0221-444444444044",
        "82293304-7118-9220-1332-555555555045",
        "93304415-8229-0331-2443-666666666046",
        "04415526-9330-1442-3554-777777777047",
        "15526637-0441-2553-4665-888888888048",
        "26637748-1552-3664-5776-999999999049",
        "37748859-2663-4775-6887-000000000050"
    ]
    return random.choice(ids)

def random_fingerprint():
    fingerprints = [
        "samsung/a54/a54:13/TP1A.220624.014/A546EXXU1AWF2:user/release-keys",
        "samsung/m34/m34:13/TP1A.220624.014/M346BXXU1AWG3:user/release-keys",
        "samsung/s23ultra/s23ultra:14/UQ1A.240205.004/S918BXXU1AXBA:user/release-keys",
        "samsung/fold5/fold5:14/UQ1A.240205.004/F946BXXU1AWM7:user/release-keys",
        "xiaomi/umi/umi:12/RKQ1.211001.001/V12.5.6.0.RJBCNXM:user/release-keys",
        "xiaomi/poco/poco:13/TKQ1.221013.002/V14.0.2.0.TKCMIXM:user/release-keys",
        "xiaomi/redmi/redmi:14/UQ1A.240205.004/V14.0.5.0.ULOMIXM:user/release-keys",
        "xiaomi/note12/note12:13/TP1A.220624.014/V14.0.1.0.TKOMIXM:user/release-keys",
        "oneplus/CPH2513/CPH2513:14/UQ1A.240205.004/EX01:user/release-keys",
        "oneplus/CPH2451/CPH2451:13/TP1A.220905.001/EX02:user/release-keys",
        "oneplus/CPH2581/CPH2581:14/UQ1A.240205.004/EX03:user/release-keys",
        "oppo/CPH2207/CPH2207:12/SKQ1.211019.001/OP01:user/release-keys",
        "oppo/CPH2419/CPH2419:13/TP1A.220624.014/OP02:user/release-keys",
        "oppo/CPH2521/CPH2521:14/UQ1A.240205.004/OP03:user/release-keys",
        "vivo/V2203/V2203:12/SP1A.210812.016/PD2203F_EX_A_12.0.10.5:user/release-keys",
        "vivo/V2254/V2254:13/TP1A.220905.001/PD2254F_EX_A_13.1.5.7:user/release-keys",
        "vivo/V2313A/V2313A:14/UQ1A.240205.004/PD2313A_EX_A_14.0.3.2:user/release-keys",
        "realme/RMX3085/RMX3085:12/SP1A.210812.016/RMX3085_11_A.24:user/release-keys",
        "realme/RMX3612/RMX3612:13/TP1A.220624.014/RMX3612_13_A.21:user/release-keys",
        "realme/RMX3491/RMX3491:14/UQ1A.240205.004/RMX3491_14_A.11:user/release-keys",
        "huawei/ANE-LX2/ANE-LX2:10/HUAWEIANE-LX2/345(user)/release-keys",
        "huawei/CDY-NX9B/CDY-NX9B:11/HUAWEICDY-NX9B/678(user)/release-keys",
        "huawei/ELS-NX9/ELS-NX9:12/HUAWEIELS-NX9/901(user)/release-keys",
        "motorola/XT2345-4/XT2345-4:13/TP1A.220624.014/20240403:user/release-keys",
        "motorola/XT2303-2/XT2303-2:14/UQ1A.240205.004/20240501:user/release-keys",
        "infinix/X6815B/X6815B:12/SP1A.210812.016/X6815B-GL-220822V123:user/release-keys",
        "infinix/X676C/X676C:13/TP1A.220624.014/X676C-H6120ABC-S-231015V104:user/release-keys",
        "tecno/CK7n/CK7n:14/UQ1A.240205.004/CK7n-H6121ABC-R-240305V103:user/release-keys",
        "tecno/CH9n/CH9n:13/TP1A.220624.014/CH9n-H6211ABC-R-231215V101:user/release-keys",
        "tecno/BD4h/BD4h:12/SP1A.210812.016/BD4h-H6112ABC-S-220915V102:user/release-keys",
        "honor/ANY-AN00/ANY-AN00:12/HONORANY-AN00/234(user)/release-keys",
        "honor/MGA-AN00/MGA-AN00:13/TP1A.220624.014/HONORMGA-AN00/567(user)/release-keys",
        "honor/LRA-AN00/LRA-AN00:14/UQ1A.240205.004/HONORLRA-AN00/890(user)/release-keys",
        "lenovo/L78051/L78051:12/SP1A.210812.016/L78051_USR_S_12.5.3:user/release-keys",
        "lenovo/K13-Note/K13-Note:13/TP1A.220624.014/K13Note_S_13.0.4:user/release-keys",
        "google/pixel7/pixel7:14/UQ1A.240205.004/10000001:user/release-keys",
        "google/pixel6a/pixel6a:14/UQ1A.240205.004/10000002:user/release-keys",
        "google/pixel5/pixel5:13/TP1A.220624.014/10000003:user/release-keys",
        "samsung/a146p/a146p:13/TP1A.220624.014/A146PXXU1AWF3:user/release-keys",
        "samsung/m54/m54:14/UQ1A.240205.004/M546BXXU1AXD2:user/release-keys",
        "xiaomi/2312DRA50G/2312DRA50G:14/UQ1A.240205.004/V14.0.7.0.UNOMIXM:user/release-keys",
        "xiaomi/23049PCD8G/23049PCD8G:13/TP1A.220624.014/V14.0.3.0.TMOMIXM:user/release-keys",
        "oneplus/CPH2459/CPH2459:14/UQ1A.240205.004/EX04:user/release-keys",
        "vivo/V2140/V2140:12/SP1A.210812.016/PD2140F_EX_A_12.0.9.8:user/release-keys",
        "realme/RMX3761/RMX3761:14/UQ1A.240205.004/RMX3761_14_A.13:user/release-keys",
        "motorola/Moto-G73/Moto-G73:13/TP1A.220624.014/20240401:user/release-keys",
        "infinix/X6711/X6711:14/UQ1A.240205.004/X6711-GL-240104V101:user/release-keys"
    ]
    return random.choice(fingerprints)

ua = [
    "Mozilla/5.0 (Linux; Android 10; SM-G960U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; P30 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; SM-A525F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Moto G Power) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Redmi Note 9S) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; OnePlus 9) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; LG G8 ThinQ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Xperia 5 II) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 4a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-S901U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ASUS_Z01QD) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Vivo V2027) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Oppo A74) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Xiaomi 12) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Nokia 7.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Realme 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Infinix Note 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Tecno Camon 18) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ZTE Axon 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-A715F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Google Pixel 6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Huawei Mate 20 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; LG V60 ThinQ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Samsung Galaxy A32) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Sony Xperia 1 III) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Google Pixel 3a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; OnePlus 8T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Redmi K40 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Xiaomi 11 Lite 5G NE) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Moto G7 Power) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-G973U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 5a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-A536B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; P40 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Redmi Note 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; OnePlus Nord 2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; LG Wing) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Xperia 1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Asus ROG Phone 3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Vivo X70 Pro+) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Oppo Reno6 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Nokia X20) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Realme 8 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Infinix Zero X Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Tecno Pova 3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ZTE Blade V2020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-A908B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Google Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; SM-G990U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]"
]
def generate_email():
    """Gumawa ng random username para sa harakirimail at bumalik ang email address at url."""
    rchjtrchjb = ''.join(random.choices(string.ascii_lowercase + string.digits, k=12))
    email_address = f"{rchjtrchjb}@harakirimail.com"
    drtyghbj5hgcbv = f"https://harakirimail.com/inbox/{rchjtrchjb}"
    return email_address, drtyghbj5hgcbv

def save_to_txt(filename, data):
    try:
        with open(filename, "a", encoding="utf-8") as f:
            f.write("|".join(data) + "\n")
    except Exception as e:
        print(f"\033[1;91m‚ùó Error saving to {filename}: {e}\033[0m")

def save_to_xlsx(filename, data):
    while True:
        with xlsx_lock:
            try:
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["NAME", "USERNAME", "PASSWORD", "ACCOUNT LINK"])
                ws.append(data)
                wb.save(filename)
                break
            except Exception as e:
                print(f"\033[1;91m‚ùó Error saving to {filename}: {e}. Retrying...\033[0m")
                time.sleep(RETRY_DELAY)

MAX_RETRIES = 3
RETRY_DELAY = 2
SUCCESS = "‚úÖ"
FAILURE = "‚ùå"
INFO = "‚ÑπÔ∏è"
WARNING = "‚ö†Ô∏è"
LOADING = "‚è≥"

def load_names_from_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file.readlines() if line.strip()]
    except FileNotFoundError:
        print(f"\033[1;91m‚ùó Error: Name file not found at {file_path}. Please create it.\033[0m")
        return ["John", "Jane", "Doe", "Smith"]

def get_names(account_type, gender):
    male_first_names_file = "first_name.txt"
    last_names_file = "last_name.txt"

    male_first_names = load_names_from_file(male_first_names_file)
    last_names = load_names_from_file(last_names_file)

    if not male_first_names:
        male_first_names = ["Juan", "Pedro"]
    if not last_names:
        last_names = ["Dela Cruz", "Reyes"]

    firstname = random.choice(male_first_names)
    lastname = random.choice(last_names)
    return firstname, lastname

def generate_random_phone_number():
    random_number = str(random.randint(1000000, 9999999))
    third = random.randint(0, 4)
    forth = random.randint(1, 7)
    return f"9{third}{forth}{random_number}"

def generate_random_password():
    base = "Promises"
    six_digit = str(random.randint(100000, 999999))
    return base + six_digit

def generate_user_details(account_type, gender, password=None):
    firstname, lastname = get_names(account_type, gender)
    year = random.randint(1978, 2001)
    date = random.randint(1, 28)
    month = random.randint(1, 12)
    if password is None:
        password = generate_random_password()
    phone_number = generate_random_phone_number()
    return firstname, lastname, date, year, month, phone_number, password

custom_password_base = None

def extract_form_details(html_content, base_url):
    """
    Extracts form action and input fields from HTML content using regex.
    This is a simplified approach and might need adjustment based on
    Facebook's actual HTML structure.
    """
    form_action_match = re.search(r'<form[^>]+action=["\']([^"\']+)["\']', html_content)
    raw_action_url = form_action_match.group(1) if form_action_match else None

    # Crucially, join the relative URL with the base URL
    action_url = urljoin(base_url, raw_action_url) if raw_action_url else None


    inputs = {}
    # Regex to find input tags and extract name and value attributes
    # Updated regex to also capture type attribute for better input identification (optional but good practice)
    input_matches = re.finditer(r'<input[^>]+name=["\']([^"\']+)["\'](?:[^>]+value=["\']([^"\']*)["\'])?', html_content)
    for match in input_matches:
        name = match.group(1)
        value = match.group(2) if match.group(2) is not None else "" # Handle cases where value attribute might be missing
        inputs[name] = value

    return action_url, inputs

def create_fbunconfirmed(account_num, account_type, gender, password=None, session=None):
    agent = random.choice(ua)
    global custom_password_base
    os.system("clear")
    email_address, drtyghbj5hgcbv = generate_email()
    if password is None:
        if custom_password_base:
            six_digit = str(random.randint(100000, 999999))
            password = custom_password_base + six_digit
        else:
            password = generate_random_password()

    firstname, lastname, date, year, month, phone_number, used_password = generate_user_details(account_type, gender, password)

    def get_page_content(url, headers, current_session):
        retries = 0
        while retries < MAX_RETRIES:
            try:
                response = current_session.get(url, timeout=30, headers=headers)
                response.raise_for_status() # Raise an exception for HTTP errors
                return response.text
            except requests.exceptions.RequestException as e:
                print(f"{FAILURE} No internet or connection issue: {e}. Retrying in {RETRY_DELAY} seconds... (Account #{account_num})")
            except Exception as e:
                print(f"{FAILURE} An unexpected error occurred: {e}. Retrying in {RETRY_DELAY} seconds... (Account #{account_num})")

            time.sleep(RETRY_DELAY)
            retries += 1
        print(f"{FAILURE} Failed to load page after {MAX_RETRIES} retries. (Account #{account_num})")
        return None

    url = "https://m.facebook.com/reg" # Base URL for registration
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://m.facebook.com/reg",
        "Connection": "keep-alive",
        "Accept-Language": "en-US,en;q=0.9",
        "X-FB-Connection-Type": "mobile.LTE",
        "X-FB-Device": random_device_model(),
        "X-FB-Device-ID": random_device_id(),
        "X-FB-Fingerprint": random_fingerprint(),
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        'x-fb-connection-type': 'Unknown',
        'accept-encoding': 'gzip, deflate',
        'content-type': 'application/x-www-form-urlencoded',
        'x-fb-http-engine': 'Liger',
        'User-Agent': agent,
    }
    if session is None:
        session = requests.Session()

    html_content = get_page_content(url, headers, session)
    if not html_content:
        print(f"\033[1;91m{FAILURE} Could not load registration page. Aborting attempt for account #{account_num}.\033[0m")
        return "FAILED_PAGE_LOAD"

    # Pass the base URL to extract_form_details
    action_url, form_inputs = extract_form_details(html_content, url)

    if not action_url:
        print(f"\033[1;91m{FAILURE} Could not find form action URL. Aborting attempt for account #{account_num}.\033[0m")
        return "FAILED_FORM_PARSE"

    # Fill in the known data
    data = {
        "firstname": firstname,
        "lastname": lastname,
        "birthday_day": str(date),
        "birthday_month": str(month),
        "birthday_year": str(year),
        "reg_email__": email_address,
        "sex": str(gender),
        "encpass": used_password,
        "submit": "Sign Up"
    }

    # Add any hidden inputs from the extracted form_inputs
    for name, value in form_inputs.items():
        if name not in data:
            data[name] = value

    try:
        response = session.post(action_url, headers=headers, data=data, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"\033[1;91m{FAILURE} Network error during submission: {e}. Cannot complete account creation for account #{account_num}.\033[0m")
        return "FAILED_SUBMISSION_NETWORK"
    except Exception as e:
        print(f"\033[1;91m{FAILURE} An unexpected error occurred during submission: {e}. Cannot complete account creation for account #{account_num}.\033[0m")
        return "FAILED_SUBMISSION_UNEXPECTED"

    with open(f"status.html", "w", encoding="utf-8") as file:
        file.write(response.text)

    if "c_user" not in session.cookies:
        print(f"\033[1;91m‚ö†Ô∏è Create Account Failed for account #{account_num}. No c_user cookie found. Try toggling airplane mode or use another email.\033[0m")
        time.sleep(3)
        return "FAILED_NO_C_USER"

    uid = session.cookies.get("c_user")
    profile_id = f'https://www.facebook.com/profile.php?id={uid}'

    cookie_dir = "/storage/emulated/0/cookie"
    os.makedirs(cookie_dir, exist_ok=True)
    cookie_file = os.path.join(cookie_dir, f"{uid}.json")
    cookie_names = ["c_user", "datr", "fr", "noscript", "sb", "xs"]
    cookies_data = {name: session.cookies.get(name, "") for name in cookie_names}
    try:
        with open(cookie_file, "w") as f:
            json.dump(cookies_data, f, indent=4)
    except IOError as e:
        print(f"\033[1;91m{FAILURE} Error saving cookies for account #{account_num}: {e}\033[0m")

    # Check for checkpoint: simplified string check
    if 'checkpoint' in response.text:
        print(f"\033[1;91m‚ö†Ô∏è Created account #{account_num} blocked. Try toggling airplane mode or clearing Facebook Lite data.\033[0m")
        time.sleep(3)
        return "BLOCKED"

    jbkj = None
    retries = 0
    while retries < MAX_RETRIES * 5:
        try:
            dtryvghjuijhn = requests.get(drtyghbj5hgcbv, timeout=30)
            dtryvghjuijhn.raise_for_status()
            # Directly search for the confirmation code in the response text
            # This assumes the confirmation code is in a predictable format within the email body
            # Example: "Your confirmation code is 123456"
            match = re.search(r'(\d+) is your confirmation code', dtryvghjuijhn.text)
            if match:
                jbkj = match.group(1)
                break
        except requests.exceptions.RequestException as e:
            print(f"\033[1;91m{FAILURE} Error fetching email for account #{account_num}: {e}. Retrying...\033[0m")
        except Exception as e:
            print(f"\033[1;91m{FAILURE} An unexpected error occurred while processing email for account #{account_num}: {e}. Retrying...\033[0m")

        time.sleep(5)
        retries += 1

    if not jbkj:
        print(f"\033[1;91m{FAILURE} Failed to get confirmation code for account #{account_num} after multiple attempts. Account might be unconfirmed.\033[0m")

    full_name = f"{firstname} {lastname}"
    data_to_save = [full_name, email_address, password, profile_id]

    with console_lock:
        print("\n\033[1;96m======================================\033[0m")
        print(f"\033[1;92m‚úÖ     Full Name: | {full_name} |\033[0m")
        print(f"\033[1;92m‚úÖ     Email: | {email_address} |\033[0m")
        print(f"\033[1;92m‚úÖ     Pass:  | {password} |\033[0m")
        print(f"\033[1;92m‚úÖ     Profile ID:  | {uid} |\033[0m")
        print(f"\033[1;92m‚úÖ     Code:  | {jbkj if jbkj else 'N/A (Code not found)'}\033[0m")
        print("\033[1;96m======================================\033[0m\n")

        while True:
            choice = input("üíæ Do you want to save this account? (y/n): ").strip().lower()
            if choice in ["y", "n"]:
                break
            print("‚ùó Please enter y or n.")

        if choice == "y":
            filename_xlsx = "/storage/emulated/0/Acc_Created.xlsx"
            filename_txt = "/storage/emulated/0/Acc_created.txt"
            save_to_xlsx(filename_xlsx, data_to_save)
            save_to_txt(filename_txt, data_to_save)
        else:
            print("\033[1;93m‚ö†Ô∏è  Account not saved.\033[0m")

    return "SUCCESS"


def NEMAIN():
    os.system("clear")
    print("\033[1;36m======================================\033[0m")
    print("\033[1;36m  Facebook By: Dars Account Creator\033[0m")
    print("\033[1;36m        (Parallel Edition)            \033[0m")
    print("\033[1;36m======================================\033[0m\n")

    global custom_password_base
    if custom_password_base is None:
        inp = 'Promises@'
        custom_password_base = inp if inp else "Promises@"

    while True:
        try:
            max_create_input = input("Enter the maximum number of accounts to create: ").strip()
            max_create = int(max_create_input)
            if max_create <= 0:
                print("\033[1;91m‚ùó Please enter a positive number.\033[0m")
            else:
                break
        except ValueError:
            print("\033[1;91m‚ùó Invalid input. Please enter a number.\033[0m")

    account_type = 1
    gender = 1

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_create) as executor:
        futures = []
        for i in range(max_create):
            time.sleep(3)
            future = executor.submit(create_fbunconfirmed, i + 1, account_type, gender, None, requests.Session())
            futures.append(future)

        for future in concurrent.futures.as_completed(futures):
            try:
                result = future.result()
                if result == "SUCCESS":
                    pass
                else:
                    print(f"\033[1;91m{WARNING} An account creation task finished with status: {result}.\033[0m")
            except Exception as exc:
                print(f"\033[1;91m{FAILURE} An account generation task generated an exception: {exc}\033[0m")

    print("\n\033[1;92m======================================\033[0m")
    print("\033[1;92m        Account Creation Finished!      \033[0m")
    print("\033[1;92m======================================\033[0m")

if __name__ == "__main__":
    NEMAIN()
