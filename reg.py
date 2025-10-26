import requests, random, json, hashlib, uuid, time
import os
import atexit
from requests.utils import dict_from_cookiejar
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from zipfile import BadZipFile
import sys

# --- ANSI escape codes for colors ---
COLOR_GREEN = '\033[92m'
COLOR_RED = '\033[91m'
COLOR_YELLOW = '\033[93m'
COLOR_CYAN = '\033[96m'
COLOR_BLUE = '\033[94m'
COLOR_RESET = '\033[0m'

# --- Global Configurations ---
COOKIE_DIR = "/storage/emulated/0/cookie"
CONFIG_FILE = "/storage/emulated/0/settings.json"
custom_password_base = None


# --- Utility Functions ---

def clear_console():
    """Clears the console screen."""
    try:
        os.system("cls" if os.name == "nt" else "clear")
    except:
        pass


def random_device_model():
    """Returns a random device model string."""
    models = [
        "Samsung-SM-S918B", "Xiaomi-2210132G", "OnePlus-CPH2451", "OPPO-CPH2207",
        "vivo-V2203", "realme-RMX3085", "Samsung-Galaxy-A54", "Samsung-SM-A146P",
        "Samsung-Galaxy-S23Ultra", "Samsung-SM-F946B", "Samsung-Galaxy-M34",
        "Xiaomi-23049PCD8G", "Xiaomi-Redmi-Note-12", "Xiaomi-POCO-X5Pro",
        "Xiaomi-2312DRA50G", "OnePlus-CPH2513", "OnePlus-CPH2581",
        "OnePlus-CPH2459", "OPPO-CPH2339", "OPPO-CPH2419", "OPPO-CPH2521",
        "vivo-V2140", "vivo-V2254", "vivo-V2230", "vivo-V2313A",
        "realme-RMX3612", "realme-RMX3571", "realme-RMX3761",
        "realme-RMX3491", "Huawei-ANE-LX2", "Huawei-JNY-LX1",
        "Huawei-ELS-NX9", "Huawei-CDY-NX9B", "Motorola-Moto-G73",
        "Motorola-XT2345-4", "Motorola-XT2303-2", "Infinix-X6815B",
        "Infinix-X6711", "Infinix-X676C", "TECNO-CK7n", "TECNO-CH9n",
        "TECNO-BD4h", "HONOR-ANY-AN00", "HONOR-MGA-AN00",
        "HONOR-LRA-AN00", "Lenovo-L78051", "Lenovo-K13-Note",
        "Google-Pixel-7", "Google-Pixel-6a", "Google-Pixel-5"
    ]
    return random.choice(models)


def random_device_id():
    """Returns a random device ID string."""
    ids = [
        "0f47e6d2-bb61-4bfc-80db-123456789001", "1a2b3c4d-5e6f-7a8b-9c0d-234567890002",
        "2b3c4d5e-6f7a-8b9c-0d1e-345678900003", "3c4d5e6f-7a8b-9c0d-1e2f-456789000004",
        "4d5e6f7a-8b9c-0d1e-2f3a-567890000005", "5e6f7a8b-9c0d-1e2f-3a4b-678900000006",
        "6f7a8b9c-0d1e-2f3a-4b5c-789000000007", "7a8b9c0d-1e2f-3a4b-5c6d-890000000008",
        "8b9c0d1e-2f3a-4b5c-6d7e-900000000009", "9c0d1e2f-3a4b-5c6d-7e8f-000000000010",
        "aa1bb2cc-3dd4-5ee6-7ff8-111111111011", "bb2cc3dd-4ee5-6ff7-8009-222222222012",
        "cc3dd4ee-5ff6-7008-9110-333333333013", "dd4ee5ff-6007-8119-0221-444444444014",
        "ee5ff600-7118-9220-1332-555555555015", "ff600711-8229-0331-2443-666666666016",
        "00611722-9330-1442-3554-777777777017", "11722833-0441-2553-4665-888888888018",
        "22833944-1552-3664-5776-999999999019", "33944a55-2663-4775-6887-000000000020",
        "44a55b66-3774-5886-7998-111111111021", "55b66c77-4885-6997-8009-222222222022",
        "66c77d88-5996-7008-9110-333333333023", "77d88e99-6007-8119-0221-444444444024",
        "88e990aa-7118-9220-1332-555555555025", "990aa1bb-8229-0331-2443-666666666026",
        "0aa1bb2c-9330-1442-3554-777777777027", "1bb2cc3d-0441-2553-4665-888888888028",
        "2cc3dd4e-1552-3664-5776-999999999029", "3dd4ee5f-2663-4775-6887-000000000030",
        "4ee5ff60-3774-5886-7998-111111111031", "5ff60071-4885-6997-8009-222222222032",
        "60071182-5996-7008-9110-333333333033", "71182293-6007-8119-0221-444444444034",
        "82293304-7118-9220-1332-555555555035", "93304415-8229-0331-2443-666666666036",
        "04415526-9330-1442-3554-777777777037", "15526637-0441-2553-4665-888888888038",
        "26637748-1552-3664-5776-999999999039", "37748859-2663-4775-6887-000000000040",
        "48859960-3774-5886-7998-111111111041", "59960071-4885-6997-8009-222222222042",
        "60071182-5996-7008-9110-333333333043", "71182293-6007-8119-0221-444444444044",
        "82293304-7118-9220-1332-555555555045", "93304415-8229-0331-2443-666666666046",
        "04415526-9330-1442-3554-777777777047", "15526637-0441-2553-4665-888888888048",
        "26637748-1552-3664-5776-999999999049", "37748859-2663-4775-6887-000000000050"
    ]
    return random.choice(ids)


def random_fingerprint():
    """Returns a random device fingerprint string."""
    fingerprints = [
        "samsung/a54/a54:13/TP1A.220624.014/A546EXXU1AWF2:user/release-keys",
        "samsung/m34/m34:13/TP1A.220624.014/M346BXXU1AWG3:user/release-keys",
        "samsung/s23ultra/s23ultra:14/UQ1A.240205.004/S918BXXU1AXBA:user/release-keys",
        "samsung/fold5/fold5:14/UQ1A.240205.004/F946BXXU1AWM7:user/release-keys",
        "xiaomi/umi/umi:12/RKQ1.211001.001/V12.5.6.0.RJBCNXM:user/release-keys",
        "xiaomi/poco/poco:13/TKQ1.221013.002/V14.0.2.0.TKCMIXM:user/release-keys",
        "xiaomi/redmi/redmi:14/UQ1A.240205.004/V14.0.5.0.ULOMIXM:user/release-keys",
        "xiaomi/note12/note12:13/TP1A.220624.014/V14.0.1.0.TKOMIXM:user/release-keys",
        "oneplus/CPH2513/CPH2513:14/UQ1A.240205.004/EX01:user/release-keys",
        "oneplus/CPH2451/CPH2451:13/TP1A.220905.001/EX02:user/release-keys",
        "oneplus/CPH2581/CPH2581:14/UQ1A.240205.004/EX03:user/release-keys",
        "oppo/CPH2207/CPH2207:12/SKQ1.211019.001/OP01:user/release-keys",
        "oppo/CPH2419/CPH2419:13/TP1A.220624.014/OP02:user/release-keys",
        "oppo/CPH2521/CPH2521:14/UQ1A.240205.004/OP03:user/release-keys",
        "vivo/V2203/V2203:12/SP1A.210812.016/PD2203F_EX_A_12.0.10.5:user/release-keys",
        "vivo/V2254/V2254:13/TP1A.220905.001/PD2254F_EX_A_13.1.5.7:user/release-keys",
        "vivo/V2313A/V2313A:14/UQ1A.240205.004/PD2313A_EX_A_14.0.3.2:user/release-keys",
        "realme/RMX3085/RMX3085:12/SP1A.210812.016/RMX3085_11_A.24:user/release-keys",
        "realme/RMX3612/RMX3612:13/TP1A.220624.014/RMX3612_13_A.21:user/release-keys",
        "realme/RMX3491/RMX3491:14/UQ1A.240205.004/RMX3491_14_A.11:user/release-keys",
        "huawei/ANE-LX2/ANE-LX2:10/HUAWEIANE-LX2/345(user)/release-keys",
        "huawei/CDY-NX9B/CDY-NX9B:11/HUAWEICDY-NX9B/678(user)/release-keys",
        "huawei/ELS-NX9/ELS-NX9:12/HUAWEIELS-NX9/901(user)/release-keys",
        "motorola/XT2345-4/XT2345-4:13/TP1A.220624.014/20240403:user/release-keys",
        "motorola/XT2303-2/XT2303-2:14/UQ1A.240205.004/20240501:user/release-keys",
        "infinix/X6815B/X6815B:12/SP1A.210812.016/X6815B-GL-220822V123:user/release-keys",
        "infinix/X676C/X676C:13/TP1A.220624.014/X676C-H6120ABC-S-231015V104:user/release-keys",
        "tecno/CK7n/CK7n:14/UQ1A.240205.004/CK7n-H6121ABC-R-240305V103:user/release-keys",
        "tecno/CH9n/CH9n:13/TP1A.220624.014/CH9n-H6211ABC-R-231215V101:user/release-keys",
        "tecno/BD4h/BD4h:12/SP1A.210812.016/BD4h-H6112ABC-S-220915V102:user/release-keys",
        "honor/ANY-AN00/ANY-AN00:12/HONORANY-AN00/234(user)/release-keys",
        "honor/MGA-AN00/MGA-AN00:13/TP1A.220624.014/HONORMGA-AN00/567(user)/release-keys",
        "honor/LRA-AN00/LRA-AN00:14/UQ1A.240205.004/HONORLRA-AN00/890(user)/release-keys",
        "lenovo/L78051/L78051:12/SP1A.210812.016/L78051_USR_S_12.5.3:user/release-keys",
        "lenovo/K13-Note/K13-Note:13/TP1A.220624.014/K13Note_S_13.0.4:user/release-keys",
        "google/pixel7/pixel7:14/UQ1A.240205.004/10000001:user/release-keys",
        "google/pixel6a/pixel6a:14/UQ1A.240205.004/10000002:user/release-keys",
        "google/pixel5/pixel5:13/TP1A.220624.014/10000003:user/release-keys",
        "samsung/a146p/a146p:13/TP1A.220624.014/A146PXXU1AWF3:user/release-keys",
        "samsung/m54/m54:14/UQ1A.240205.004/M546BXXU1AXD2:user/release-keys",
        "xiaomi/2312DRA50G/2312DRA50G:14/UQ1A.240205.004/V14.0.7.0.UNOMIXM:user/release-keys",
        "xiaomi/23049PCD8G/23049PCD8G:13/TP1A.220624.014/V14.0.3.0.TMOMIXM:user/release-keys",
        "oneplus/CPH2459/CPH2459:14/UQ1A.240205.004/EX04:user/release-keys",
        "vivo/V2140/V2140:12/SP1A.210812.016/PD2140F_EX_A_12.0.9.8:user/release-keys",
        "realme/RMX3761/RMX3761:14/UQ1A.240205.004/RMX3761_14_A.13:user/release-keys",
        "motorola/Moto-G73/Moto-G73:13/TP1A.220624.014/20240401:user/release-keys",
        "infinix/X6711/X6711:14/UQ1A.240205.004/X6711-GL-240104V101:user/release-keys"
    ]
    return random.choice(fingerprints)


# User Agents from reg.py (omitted for brevity)
ua = [
    # ...
]


def delete_config_file():
    """Clears the console screen."""
    # NO LONGER DELETES THE CONFIG FILE ON EXIT
    pass


atexit.register(delete_config_file)  # Register a dummy function to avoid error


def save_user_choice(key, value):
    """Saves user preferences to the config file."""
    data = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except:
                data = {}
    data[key] = value
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def load_user_choice(key):
    """Loads user preferences from the config file."""
    if not os.path.exists(CONFIG_FILE):
        return None
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            return data.get(key)
        except:
            return None


def save_to_txt(filename, data):
    """Saves account data to a TXT file."""
    try:
        with open(filename, "a", encoding="utf-8") as f:
            f.write("|".join(data) + "\n")
    except Exception as e:
        print(f"\033[1;91m❗ Error saving to {filename}: {e}\033[0m")


def save_to_xlsx(filename, data):
    """Saves account data to an XLSX file, checking for existing entries."""
    # MODIFICATION: Added 'ACCESS TOKEN' to the header
    header_columns = ['NAME', 'USERNAME', 'PASSWORD', 'ACCOUNT LINK', 'ACCESS TOKEN']

    while True:
        try:
            if os.path.exists(filename):
                try:
                    wb = load_workbook(filename)
                    ws = wb.active
                except BadZipFile:
                    print(f"\033[91m⚠️ Corrupted XLSX detected at {filename}. Recreating file...\033[0m")
                    os.remove(filename)
                    wb = Workbook()
                    ws = wb.active
                    ws.append(header_columns)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(header_columns)

            header = [cell.value for cell in ws[1]]
            if header != header_columns:
                ws.delete_rows(1)
                ws.insert_rows(1)
                ws.append(header_columns)

            existing_rows = [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]
            if tuple(data) not in existing_rows:
                ws.append(data)

            wb.save(filename)
            break
        except Exception as e:
            print(f"❗ Error saving to {filename}: {e}. Retrying in 1 second...")
            time.sleep(1)


def load_names_from_file(file_path):
    """Loads a list of names from a given file path."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file if line.strip()]
    except FileNotFoundError:
        print(f"{COLOR_RED}❌ ERROR: Name file not found: {file_path}{COLOR_RESET}")
        return ["Unknown"]


def get_names(account_type, gender):
    """Gets a random first and last name from name files."""
    firstnames = load_names_from_file("first_name.txt")
    last_names = load_names_from_file("last_name.txt")
    firstname = random.choice(firstnames)
    lastname = random.choice(last_names)
    return firstname, lastname


def generate_random_phone_number():
    """Generates a random phone number in a specific format."""
    random_number = str(random.randint(1000000, 9999999))
    third = random.randint(0, 4)
    forth = random.randint(1, 7)
    return f"9{third}{forth}{random_number}"


def generate_random_password():
    """Generates a random password."""
    return 'Promises' + str(random.randint(100000, 999999))


def generate_user_details(account_type, gender, password=None):
    """Generates a full set of user details for registration."""
    firstname, lastname = get_names(account_type, gender)
    year = random.randint(1978, 2001)
    date = random.randint(1, 28)
    month = random.randint(1, 12)
    if password is None:
        password = generate_random_password()
    phone_number = generate_random_phone_number()
    return firstname, lastname, date, year, month, phone_number, password


# --- Login Function (Modified with Max Retries and better logging) ---
def Login(email: str, password: str, max_retries=3):
    """
    Attempts to login to Facebook's b-graph API with up to 3 retries
    to fetch the Access Token.
    """
    r = requests.Session()
    attempt = 0

    # Define headers
    head = {
        'Host': 'b-graph.facebook.com',
        'X-Fb-Connection-Quality': 'EXCELLENT',
        'Authorization': 'OAuth 350685531728|62f8ce9f74b12f84c123cc23437a4a32',
        'User-Agent': 'Dalvik/2.1.0 (Linux; U; Android 7.1.2; RMX3740 Build/QP1A.190711.020) [FBAN/FB4A;FBAV/417.0.0.33.65;FBPN/com.facebook.katana;FBLC/in_ID;FBBV/480086274;FBCR/Corporation Tbk;FBMF/realme;FBBD/realme;FBDV/RMX3740;FBSV/7.1.2;FBCA/x86:armeabi-v7a;FBDM/{density=1.0,width=540,height=960};FB_FW/1;FBRV/483172840;]',
        'X-Tigon-Is-Retry': 'false',
        'X-Fb-Friendly-Name': 'authenticate',
        'X-Fb-Connection-Bandwidth': str(random.randrange(70000000, 80000000)),
        'Zero-Rated': '0',
        'X-Fb-Net-Hni': str(random.randrange(50000, 60000)),
        'X-Fb-Sim-Hni': str(random.randrange(50000, 60000)),
        'X-Fb-Request-Analytics-Tags': '{"network_tags":{"product":"350685531728","retry_attempt":"0"},"application_tags":"unknown"}',
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-Fb-Connection-Type': 'WIFI',
        'X-Fb-Device-Group': str(random.randrange(4700, 5000)),
        'Priority': 'u=3,i',
        'Accept-Encoding': 'gzip, deflate',
        'X-Fb-Http-Engine': 'Liger',
        'X-Fb-Client-Ip': 'true',
        'X-Fb-Server-Cluster': 'true',
        'Content-Length': str(random.randrange(1500, 2000)),
        'cache-control': "private, no-cache, no-store, must-revalidate",
        "facebook-api-version": "v1.0",
        "pragma": "no-cache",
        "priority": "u=0,i",
        "strict-transport-security": "max-age=15552000; preload",
        "vary": "Accept-Encoding",
        "x-fb-connection-quality": "GOOD; q=0.7, rtt=73, rtx=0, c=23, mss=1232, tbw=5012, tp=10, tpl=0, uplat=405, ullat=0",
        "x-fb-debug": "g/lwUlHD6vXZly0pnMoWnhifQ8PoyIuzDnUKVk5ZWru6+2XT2yaUB9Y/TSXbt0/637lElrllnUhGyXNJLheBKA==",
        "x-fb-request-id": "AEJauAi2IHwyhd_zl3pC-4E",
        "x-fb-rev": "1025308755",
        "x-fb-trace-id": "C/GnaBOOeUa",
        "x-frame-options": "DENY"
    }

    # Define data payload
    data = {
        'adid': str(uuid.uuid4()),
        'format': 'json',
        'device_id': str(uuid.uuid4()),
        'email': email,
        'password': f'#PWD_FB4A:0:{str(time.time())[:10]}:{password}',
        'generate_analytics_claim': '1',
        'community_id': '',
        'linked_guest_account_userid': '',
        'cpl': True,
        'try_num': '1',
        'family_device_id': str(uuid.uuid4()),
        'secure_family_device_id': str(uuid.uuid4()),
        'credentials_type': 'password',
        'account_switcher_uids': [],
        'fb4a_shared_phone_cpl_experiment': 'fb4a_shared_phone_nonce_cpl_at_risk_v3',
        'fb4a_shared_phone_cpl_group': 'enable_v3_at_risk',
        'enroll_misauth': False,
        'generate_session_cookies': '1',
        'error_detail_type': 'button_with_disabled',
        'source': 'login',
        'machine_id': ''.join(
            [random.choice('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789') for i in range(24)]),
        'jazoest': str(random.randrange(22000, 23000)),
        'meta_inf_fbmeta': 'V2_UNTAGGED',
        'advertiser_id': str(uuid.uuid4()),
        'encrypted_msisdn': '',
        'currently_logged_in_userid': '0',
        'locale': 'id_ID',
        'client_country_code': 'ID',
        'fb_api_req_friendly_name': 'authenticate',
        'fb_api_caller_class': 'Fb4aAuthHandler',
        'api_key': '882a8490361da98702bf97a021ddc14d',
        'sig': hashlib.md5(str(uuid.uuid4()).encode()).hexdigest()[:32],
        'access_token': '350685531728|62f8ce9f74b12f84c123cc23437a4a32'
    }

    while attempt < max_retries:
        attempt += 1
        print("-" * 30)
        print(f"{COLOR_YELLOW}🔑 Token Fetch Attempt {attempt}/{max_retries}...{COLOR_RESET}")

        try:
            pos = r.post('https://b-graph.facebook.com/auth/login', data=data, headers=head, timeout=15).json()

            if ('session_key' in str(pos)) and ('access_token' in str(pos)):
                token = pos['access_token']
                print(f'{COLOR_GREEN}✅ TOKEN STATUS: SUCCESS{COLOR_RESET}')
                # MODIFICATION: Ipakita lang ang prefix na "EAAAAUa..." sa console
                print(f'{COLOR_YELLOW}TOKEN:{COLOR_RESET} EAAAAUa...{COLOR_RESET}')
                # Ang buong token ay ibabalik pa rin para ma-save
                return token, 'SUCCESS'
            else:
                if 'error' in pos and 'message' in pos['error']:
                    error_message = pos["error"]["message"]
                    print(f'{COLOR_RED}❌ LOGIN ERROR: {error_message}{COLOR_RESET}')
                    # If the error is permanent (like bad password/user), stop retrying
                    if "password" in error_message.lower() or "user" in error_message.lower() or "checkpoint" in error_message.lower():
                        return None, error_message

                    if attempt < max_retries:
                        print(f'{COLOR_YELLOW}Retrying in 5 seconds...{COLOR_RESET}')
                        time.sleep(5)
                else:
                    print(f'{COLOR_RED}❌ UNKNOWN ERROR FORMAT: {pos}{COLOR_RESET}')
                    if attempt < max_retries:
                        print(f'{COLOR_YELLOW}Retrying in 5 seconds...{COLOR_RESET}')
                        time.sleep(5)

        except requests.exceptions.RequestException as e:
            print(f'{COLOR_RED}❌ CONNECTION ERROR: {e}{COLOR_RESET}')
            if attempt < max_retries:
                print(f'{COLOR_YELLOW}Retrying in 5 seconds...{COLOR_RESET}')
                time.sleep(5)

    print("-" * 30)
    print(f"{COLOR_RED}❗ TOKEN FAILED: Could not retrieve token after {max_retries} attempts.{COLOR_RESET}")
    print("-" * 30)
    # MODIFICATION: Return None and failure status
    return None, 'MAX_RETRIES_EXCEEDED'


# --- Registration Logic (Modified for cleaner logging and no save prompt) ---
def create_fbunconfirmed(account_type, usern, gender, password=None, session=None):
    """Handles Facebook account registration and asks the user about token fetching/saving."""
    global custom_password_base

    if password is None:
        if custom_password_base:
            password = custom_password_base + str(random.randint(100000, 999999))
        else:
            password = generate_random_password()

    firstname, lastname, date, year, month, phone_number, used_password = generate_user_details(account_type, gender,
                                                                                                password)
    email_or_phone = ""
    is_phone_choice = False

    url = "https://m.facebook.com/reg"
    # Headers from reg.py
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://m.facebook.com/reg",
        "Connection": "keep-alive",
        "Accept-Language": "en-US,en;q=0.9",
        "X-FB-Connection-Type": "mobile.LTE",
        "X-FB-Device": random_device_model(),
        "X-FB-Device-ID": random_device_id(),
        "X-FB-Fingerprint": random_fingerprint(),
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        'x-fb-connection-type': 'Unknown',
        'accept-encoding': 'gzip, deflate',
        'content-type': 'application/x-www-form-urlencoded',
        'x-fb-http-engine': 'Liger',
        'User-Agent': 'Dalvik/2.1.0 (Linux; U; Android 14; T614SP Build/UP1A.231005.007) [FBAN/Orca-Android;FBAV/513.1.0.46.107;FBPN/com.facebook.orca;FBLC/en_US;FBBV/753632239;FBCR/HOME;FBMF/TCL;FBBD/TCL;FBDV/T614SP;FBSV/14;FBCA/arm64-v8a:null;FBDM/{density=2.0,width=720,height=1489};FB_FW/1;]',
    }

    if session is None:
        session = requests.Session()

    def get_registration_form():
        """Fetches the registration form to get initial inputs."""
        while True:
            try:
                print(f"{COLOR_YELLOW}⚙️ Step 1/3: Fetching registration form...{COLOR_RESET}")
                response = session.get(url, headers=headers, timeout=60)
                soup = BeautifulSoup(response.text, "html.parser")
                form = soup.find("form")
                if form:
                    print(f"{COLOR_GREEN}✅ Form fetched successfully.{COLOR_RESET}")
                    time.sleep(3)
                    clear_console()
                    return form
            except:
                print(f'{COLOR_RED}❌ Connection failed. Retrying in 3 seconds...{COLOR_RESET}')
                time.sleep(3)

    form = get_registration_form()

    # Choice input with saved preference
    choice = load_user_choice("reg_choice")

    if choice is None:
        while True:
            print("\n" + "=" * 30)
            print(f"{COLOR_BLUE}Choose Registration Method:{COLOR_RESET}")
            print(" [1] Enter Email")
            print(" [2] Use Random Phone Number")
            choice = input(f"{COLOR_GREEN}Your choice (1 or 2): {COLOR_RESET}").strip()
            if choice in ['1', '2']:
                clear_console()
                save_user_choice("reg_choice", choice)
                break
            else:
                print(f"{COLOR_RED}❌ Invalid choice. Please enter 1 or 2.{COLOR_RESET}")
    else:
        print(
            f"\n{COLOR_BLUE}Using saved choice: [{choice}] {'Enter Email' if choice == '1' else 'Use Random Phone Number'}{COLOR_RESET}")

    if choice == '1':
        while True:
            email_or_phone = input(f"{COLOR_GREEN}Paste your email:{COLOR_RESET}").strip()
            if email_or_phone:
                break
            print(f"{COLOR_RED}❌ Email cannot be empty.{COLOR_RESET}")
        is_phone_choice = False
    else:
        email_or_phone = phone_number
        print(f"{COLOR_GREEN}Generated Phone Number:{COLOR_RESET} {email_or_phone}")
        is_phone_choice = True

    print(f"{COLOR_YELLOW}⚙️ Step 2/3: Submitting registration request...{COLOR_RESET}")

    # Registration data payload
    data = {
        "firstname": firstname,
        "lastname": lastname,
        "birthday_day": str(date),
        "birthday_month": str(month),
        "birthday_year": str(year),
        "reg_email__": email_or_phone,
        "sex": str(gender),
        "encpass": used_password,
        "submit": "Sign Up"
    }

    if form:
        action_url = requests.compat.urljoin(url, form.get("action", url))
        for inp in form.find_all("input"):
            if inp.has_attr("name") and inp["name"] not in data:
                data[inp["name"]] = inp.get("value", "")

        try:
            session.post(action_url, headers=headers, data=data, timeout=60)
        except requests.exceptions.RequestException:
            pass

    final_username = email_or_phone
    registration_successful = "c_user" in session.cookies

    print("-" * 30)

    if not registration_successful:
        print(f"\n{COLOR_RED}❌ REGISTRATION FAILED: No 'c_user' cookie found after submission.{COLOR_RESET}")
        print(f"{COLOR_YELLOW}💡 Suggestion: Try toggling airplane mode or use another email.{COLOR_RESET}")
        time.sleep(3)
        return

    print(f"{COLOR_GREEN}✅ REGISTRATION SUCCESSFUL!{COLOR_RESET}")

    # Change email if generated with phone
    if is_phone_choice and registration_successful:
        print("\n" + "=" * 30)
        print(f"{COLOR_BLUE}⚙️ Phone registration detected. Changing to Email...{COLOR_RESET}")
        print("=" * 30)
        while True:
            try:
                new_email = input(f"{COLOR_GREEN}Please enter your NEW email for the account: {COLOR_RESET}").strip()
                if not new_email:
                    print(f"{COLOR_RED}❌ Email cannot be empty. Skipping email change.{COLOR_RESET}")
                    break

                change_email_url = "https://m.facebook.com/changeemail/"
                print(f"{COLOR_YELLOW}⚙️ Submitting email change to: {new_email}{COLOR_RESET}")

                response = session.get(change_email_url, headers=headers, timeout=60)
                soup = BeautifulSoup(response.text, "html.parser")
                form = soup.find("form")

                if not form:
                    print(f"{COLOR_RED}❌ Could not load email change form. Skipping.{COLOR_RESET}")
                    break

                action_url = requests.compat.urljoin(change_email_url, form.get("action", change_email_url))
                data = {}
                for inp in form.find_all("input"):
                    if inp.has_attr("name"):
                        data[inp["name"]] = inp.get("value", "")

                data["new"] = new_email
                data["submit"] = "Add"

                session.post(action_url, headers=headers, data=data, timeout=60)
                check_response = session.get(change_email_url, headers=headers, timeout=60)

                if "email" in check_response.text.lower():
                    print(f"{COLOR_GREEN}✅ Email change submitted successfully! (Needs verification){COLOR_RESET}")
                else:
                    print(f"{COLOR_YELLOW}⚠️ Email change status unclear. Check account manually.{COLOR_RESET}")

                final_username = new_email
                break
            except Exception as e:
                print(f"{COLOR_RED}❌ Error changing email: {e}{COLOR_RESET}")
                time.sleep(2)

    full_name = f"{firstname} {lastname}"
    uid = session.cookies.get("c_user")
    profile_id = f'https://www.facebook.com/profile.php?id={uid}'

    # --- New Logic: Ask to get Token ---
    final_token = 'USER_SKIPPED_TOKEN_FETCH' # Default value if user skips

    print("\n" + "=" * 50)
    print(f"{COLOR_BLUE}           ACCOUNT REGISTRATION COMPLETE{COLOR_RESET}")
    print("=" * 50)
    
    # ASK THE USER
    while True:
        get_token_choice = input(f"{COLOR_YELLOW}Do you want to proceed to Step 3: FETCH ACCESS TOKEN? (y/n): {COLOR_RESET}").strip().lower()
        if get_token_choice in ['y', 'n']:
            break
        else:
            print(f"{COLOR_RED}❌ Invalid choice. Please enter 'y' or 'n'.{COLOR_RESET}")
    
    if get_token_choice == 'y':
        print("\n" + "=" * 50)
        print(f"{COLOR_BLUE}           Step 3: FETCHING TOKEN & SAVING{COLOR_RESET}")
        print("=" * 50)

        # Call Login to get Token and status
        token, status = Login(email=final_username, password=used_password, max_retries=3)

        # Use the retrieved token, or 'FAILED_TO_GET_TOKEN' if not successful
        final_token = token if token else 'FAILED_TO_GET_TOKEN'
        
        # --- Token Retry Logic ---
        if final_token == 'FAILED_TO_GET_TOKEN':
            print("\n" + "=" * 50)
            print(f"{COLOR_RED}❗ TOKEN FETCH FAILED after initial attempts!{COLOR_RESET}")
            
            while True:
                retry_choice = input(f"{COLOR_YELLOW}Do you want to attempt re-login for the token? (y/n): {COLOR_RESET}").strip().lower()
                if retry_choice in ['y', 'n']:
                    break
                else:
                    print(f"{COLOR_RED}❌ Invalid choice. Please enter 'y' or 'n'.{COLOR_RESET}")
            
            if retry_choice == 'y':
                print("\n" + "=" * 50)
                print(f"{COLOR_BLUE}           RE-ATTEMPTING TOKEN FETCH{COLOR_RESET}")
                print("=" * 50)
                # Increased max retries for manual attempt
                token, status = Login(email=final_username, password=used_password, max_retries=5) 
                final_token = token if token else 'FAILED_TO_GET_TOKEN_RETRY'
                
                if final_token not in ['FAILED_TO_GET_TOKEN_RETRY', 'FAILED_TO_GET_TOKEN']:
                    print(f"{COLOR_GREEN}✅ TOKEN SUCCESSFULLY RETRIEVED on re-attempt!{COLOR_RESET}")
                else:
                    print(f"{COLOR_RED}❌ TOKEN STILL FAILED after re-attempt.{COLOR_RESET}")
            else:
                final_token = 'USER_SKIPPED_TOKEN_RETRY'
    
    # Automatic Saving Logic
    filename_xlsx = "/storage/emulated/0/Acc_Created.xlsx"
    filename_txt = "/storage/emulated/0/Acc_created.txt"

    # MODIFICATION: ADDED final_token TO THE data_to_save LIST
    data_to_save = [full_name, final_username, used_password, profile_id, final_token]

    save_to_xlsx(filename_xlsx, data_to_save)
    save_to_txt(filename_txt, data_to_save)

    print("\n" + "=" * 50)
    print(f"{COLOR_BLUE}           FINAL ACCOUNT DETAILS{COLOR_RESET}")
    print("=" * 50)
    print(f"{COLOR_GREEN}👤 Full Name:{COLOR_RESET} {full_name}")
    print(f"{COLOR_GREEN}📧 Username:{COLOR_RESET} {final_username}")
    print(f"{COLOR_GREEN}🔑 Password:{COLOR_RESET} {used_password}")
    print(f"{COLOR_GREEN}🔗 Profile ID:{COLOR_RESET} {profile_id}")
    
    # Conditional Token Display
    token_display = 'EAAAAUa...' if final_token.startswith('EAAAAUa') else final_token
    if final_token.startswith('EAAAAUa'):
        print(f"{COLOR_GREEN}🔐 Access Token:{COLOR_RESET} {token_display}")
    else:
        print(f"{COLOR_RED}🔐 Access Token:{COLOR_RESET} {token_display}")
        
    print(f"{COLOR_GREEN}💾 SAVE STATUS: Account saved successfully to storage.{COLOR_RESET}")
    print("=" * 50)


def NEMAIN():
    """Main registration sequence, now runs in a loop."""
    clear_console()

    account_type = 1
    gender = 1

    global custom_password_base
    if custom_password_base is None:
        inp = input(f"{COLOR_GREEN}😊 Type your password: {COLOR_RESET}").strip()
        custom_password_base = inp if inp else "Promises"

    # MODIFICATION: Use an infinite loop to keep creating accounts
    count = 0
    while True:
        count += 1
        session = requests.Session()
        print("\n" + "=" * 50)
        print(f"{COLOR_CYAN}STARTING ACCOUNT REGISTRATION NO. {count}{COLOR_RESET}")
        print("=" * 50)

        # Pass a fresh session for each attempt
        create_fbunconfirmed(account_type, "ali", gender, session=session)

        print("\n" + "=" * 50)
        print(f"{COLOR_CYAN}Completed account {count}. Preparing for the next account in 10 seconds...{COLOR_RESET}")
        print("=" * 50)
        time.sleep(10)


# --- Main Execution (Modified to run NEMAIN in a loop) ---
if __name__ == "__main__":
    # Clear old settings file on first run only
    # Note: We keep the settings file after the first successful choice for the loop
    if os.path.exists(CONFIG_FILE):
        try:
            # Check if settings.json is from a previous run or if we should keep it
            # Simple check: only delete if it's empty, otherwise keep user choice.
            if os.path.getsize(CONFIG_FILE) == 0:
                os.remove(CONFIG_FILE)
        except:
            pass

    # Run NEMAIN indefinitely
    NEMAIN()

    # The program will now loop indefinitely until manually stopped.
