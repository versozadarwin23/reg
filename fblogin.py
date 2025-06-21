import os
import requests
from bs4 import BeautifulSoup
import time
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
import re
import json
from requests.utils import dict_from_cookiejar, cookiejar_from_dict
from openpyxl import load_workbook

os.system("clear")

headers = {
    'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; CPH1903 Build/O11019; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/70.0.3538.110 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/444.0.0.0.110;]',
    'Referer': 'https://m.facebook.com/',
    'Content-Type': 'application/x-www-form-urlencoded'
}

windows_headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
    'Referer': 'https://m.facebook.com/',
    'Content-Type': 'application/x-www-form-urlencoded'
}

lock = Lock()
success_count = 0
error_count = 0
logged_accounts = set()

def keep_alive(name, username, password, account_link):
    global success_count, error_count

    session = requests.Session()
    cookie_file = f"/storage/emulated/0/cookie/{username}.json"

    reused_session = False

    if os.path.exists(cookie_file):
        try:
            with open(cookie_file, "r") as f:
                cookies = json.load(f)
            session.cookies = cookiejar_from_dict(cookies)

            home_check = session.get("https://m.facebook.com/home.php", headers=windows_headers, timeout=15)
            if "c_user" in session.cookies:
                with lock:
                    success_count += 1
                uid = session.cookies.get("c_user")
                print(f"\033[92m[✓] {name} Keep-alive OK: {uid} |  (session reused)\033[0m")
                reused_session = True
        except:
            pass

    if not reused_session:
        try:
            login_page = session.get('https://m.facebook.com/login', headers=headers, timeout=60, allow_redirects=True)
            soup = BeautifulSoup(login_page.text, 'html.parser')

            form = soup.find('form', {'id': 'login_form'})
            if not form:
                with lock:
                    error_count += 1
                return

            action = form.get('action')
            post_url = 'https://m.facebook.com' + action
            sleep_time = random.uniform(5.7, 7.3)  # gives a float between 5.0 and 7.0
            payload = {'email': username, 'pass': password}
            for input_tag in form.find_all('input'):
                time.sleep(sleep_time)
                name_input = input_tag.get('name')
                value = input_tag.get('value', '')
                if name_input and name_input not in payload:
                    payload[name_input] = value

            response = session.post(post_url, data=payload, headers=headers, timeout=60, allow_redirects=True)

            if "c_user" in session.cookies:
                uid = session.cookies.get("c_user")
                with lock:
                    success_count += 1
                os.makedirs("/storage/emulated/0/cookie", exist_ok=True)
                cookies_dict = dict_from_cookiejar(session.cookies)
                cookies_dict['active_time'] = "0m"
                with open(cookie_file, "w") as f:
                    json.dump(cookies_dict, f)
                print(f"\033[92m[✓] {name} Keep-alive OK: {uid} | (login success)\033[0m")
            else:
                with lock:
                    error_count += 1
                return
        except Exception as e:
            with lock:
                error_count += 1
            return

    start_time = time.time()
    retry_count = 0
    max_retries = 3

    while True:
        try:
            url = 'https://m.facebook.com/home.php'
            response = session.get(url, headers=windows_headers, timeout=10, allow_redirects=True)

            if "c_user" in session.cookies:
                retry_count = 0
                uid = session.cookies.get("c_user")
                elapsed_seconds = time.time() - start_time
                elapsed_minutes = int(elapsed_seconds // 60)
                hours = elapsed_minutes // 60
                minutes = elapsed_minutes % 60
                active_time = f"{hours}h {minutes}m" if hours > 0 else f"{minutes}m"

                cookies_dict = dict_from_cookiejar(session.cookies)
                cookies_dict['active_time'] = active_time
                with open(cookie_file, "w") as f:
                    json.dump(cookies_dict, f)

                print(f"\033[92m[✓] {name} Keep-alive OK: {uid} | Active: {active_time}\033[0m")
            else:
                return

        except requests.exceptions.RequestException:
            retry_count += 1
            if retry_count >= max_retries:
                start_time = time.time()
                retry_count = 0

        except:
            pass

        time.sleep(60)

def load_accounts():
    accounts = []
    pattern = re.compile(r'^https://www\.facebook\.com/profile\.php\?id=')

    filepath = "/storage/emulated/0/Acc_Created.xlsx"
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = row[0]
                email_or_id = row[1]
                password = row[2]
                account_link = row[3] if len(row) > 3 else ""

                if not account_link or not password:
                    continue
                username = pattern.sub('', account_link)
                accounts.append([name, username, password, account_link])
        except Exception as e:
            print(f"[ERROR] Failed to load from XLSX: {e}")

    if not accounts:
        print("[INFO] Falling back to cookie files...")
        cookie_dir = "/storage/emulated/0/cookie"
        if os.path.exists(cookie_dir):
            for filename in os.listdir(cookie_dir):
                if filename.endswith(".json"):
                    try:
                        username = filename.replace(".json", "")
                        with open(os.path.join(cookie_dir, filename), "r") as f:
                            cookies = json.load(f)
                        name = username
                        password = "unknown"
                        account_link = f"https://www.facebook.com/profile.php?id={username}"
                        accounts.append([name, username, password, account_link])
                    except Exception as e:
                        print(f"[WARN] Failed to read cookie {filename}: {e}")

    return accounts

def main():
    executor = None
    prev_account_count = 0

    while True:
        accounts = load_accounts()
        current_account_count = len(accounts)

        if current_account_count != prev_account_count:
            if executor:
                executor.shutdown(wait=False)
            max_workers = current_account_count if current_account_count > 0 else 1
            executor = ThreadPoolExecutor(max_workers=max_workers)
            prev_account_count = current_account_count

        for name, username, password, account_link in accounts:
            if username not in logged_accounts:
                logged_accounts.add(username)
                executor.submit(keep_alive, name, username, password, account_link)

        time.sleep(60)

if __name__ == "__main__":
    main()
