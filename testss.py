import json
import re
import os
import requests
import time
import random
from bs4 import BeautifulSoup
from requests.utils import dict_from_cookiejar, cookiejar_from_dict
from openpyxl import Workbook, load_workbook
os.system("clear")

class FacebookAccountHandler:
    """Handler for creating and keeping Facebook unconfirmed accounts alive."""

    HEADERS = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://m.facebook.com/reg",
        "Connection": "keep-alive",
        "X-FB-Connection-Type": "MOBILE.LTE",
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        'x-fb-connection-type': 'Unknown',
        'accept-encoding': 'gzip, deflate',
        'content-type': 'application/x-www-form-urlencoded',
        'x-fb-http-engine': 'Liger',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; CPH1903 Build/O11019; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/70.0.3538.110 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/444.0.0.0.110;]',
    }

    BASE_PASSWORD = "promises"

    def __init__(self, account_type=1, gender=1, session=None, save_dir="/storage/emulated/0/"):
        """Initialize the handler."""
        self.account_type = account_type
        self.gender = gender
        self.session = session or requests.Session()
        self.save_dir = save_dir
        os.makedirs(f"{save_dir}/cookie", exist_ok=True)

    def save_to_xlsx(self, filename, data):
        """Save account data to an Excel workbook."""
        while True:
            try:
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["NAME", "USERNAME", "PASSWORD", "ACCOUNT LINK"])
                ws.append(data)
                wb.save(filename)
                break
            except Exception as e:
                print(f"Error saving to {filename}: {e}. Retrying...")
                time.sleep(1)

    def load_names_from_file(self, file_path):
        """Load names from a text file."""
        with open(file_path, 'r') as file:
            return [line.strip() for line in file.readlines()]

    def get_names(self):
        """Get random first and last names based on account type and gender."""
        if self.account_type == 1:
            male_first_names = self.load_names_from_file("first_name.txt")
            last_names = self.load_names_from_file("last_name.txt")
            female_first_names = []
        else:
            male_first_names = []
            female_first_names = self.load_names_from_file("path_to_female_first_names.txt")
            last_names = self.load_names_from_file("path_to_last_names.txt")

        firstname = random.choice(male_first_names if self.gender == 1 else female_first_names)
        lastname = random.choice(last_names)
        return firstname, lastname

    def generate_user_details(self, password=None):
        """Generate user details including names, birth date, and phone number."""
        firstname, lastname = self.get_names()
        year = random.randint(1978, 2001)
        date = random.randint(1, 28)
        month = random.randint(1, 12)

        if password is None:
            password = f"{self.BASE_PASSWORD}{random.randint(100000, 999999)}"

        phone_number = f"9{random.randint(0, 4)}{random.randint(1, 7)}{random.randint(1000000, 9999999)}"
        return firstname, lastname, date, year, month, phone_number, password

    def create_account(self, email_or_phone, password=None):
        """Create an unconfirmed Facebook account."""
        firstname, lastname, date, year, month, phone_number, password = self.generate_user_details(password)
        url = "https://m.facebook.com/reg"

        # Get Registration Page
        try:
            resp = self.session.get(url, headers=self.HEADERS, timeout=10)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"Error accessing registration page: {e}")
            return

        while True:
            try:
                soup = BeautifulSoup(resp.text, "html.parser")  # Parse the response
                form = soup.find("form")  # Find the form
                if not form:
                    print("Error: Registration form not found.")
                    return
                else:
                    break  # Form found, break the loop
            except Exception as e:
                print(f"Error parsing the page: {e}")
                return

        action_url = requests.compat.urljoin(url, form.get("action", url))
        inputs = form.find_all("input")

        data = {
            "firstname": firstname,
            "lastname": lastname,
            "birthday_day": str(date),
            "birthday_month": str(month),
            "birthday_year": str(year),
            "reg_email__": email_or_phone,
            "sex": str(self.gender),
            "encpass": password,
            "submit": "Sign Up"
        }

        for inp in inputs:
            if inp.get("name") and inp.get("name") not in data:
                data[inp["name"]] = inp.get("value", "")

        try:
            resp = self.session.post(action_url, headers=self.HEADERS, data=data, allow_redirects=True, timeout=60)
        except requests.RequestException as e:
            print(f"Error submitting the registration form: {e}")
            return

        uid = self.session.cookies.get("c_user")
        profile_id = f'https://www.facebook.com/profile.php?id={uid}'

        if "c_user" not in self.session.cookies:
            print("\033[1;91m⚠️ Create Account Failed.\033[0m")
            return

        cookie_file = f'{self.save_dir}/cookie/{uid}.json'
        cookie_data = dict_from_cookiejar(self.session.cookies)

        try:
            with open(cookie_file, "w") as f:
                json.dump(cookie_data, f)
        except Exception as e:
            print(f"\033[1;91m⚠️ Failed to save cookie: {e}\033[0m")

        print(f"\033[1;92mAccount | Pass: {password}\033[0m")

        # if "checkpoint" in resp.text:
        #     print(
        #         "\033[1;91m⚠️ Created account is blocked (checkpoint required). Trying phone number instead...\033[0m")
        #     phone_only = phone_number
        #     return self.create_account(phone_only)

            # Save Account Details
        filename = f"{self.save_dir}/Acc_Created.xlsx"
        full_name = f"{firstname} {lastname}"
        self.save_to_xlsx(filename, [full_name, email_or_phone, password, profile_id])

        print(f"\033[1;92m✅ Account saved: {firstname} {lastname} | Pass: {password}\033[0m")
        return uid, profile_id, password

    def keep_alive(self, uid, password, interval=60, max_retries=3):
        """Keep account active and detect checkpoints."""
        cookie_file = f"{self.save_dir}/cookie/{uid}.json"
        if not os.path.exists(cookie_file):
            print(f"\033[1;91m⚠️ No cookie found for UID {uid}.\033[0m")
            return

        with open(cookie_file, "r") as f:
            self.session.cookies = cookiejar_from_dict(json.load(f))

        start_time = time.time()
        while True:
            try:
                resp = self.session.get("https://m.facebook.com/home.php", headers=self.HEADERS, timeout=60)

                if "c_user" in self.session.cookies:
                    elapsed_minutes = int((time.time() - start_time) / 60)
                    hours = elapsed_minutes // 60
                    mins = elapsed_minutes % 60
                    active_time = f"{hours}h {mins}m" if hours > 0 else f"{mins}m"

                    cookie_data = dict_from_cookiejar(self.session.cookies)
                    cookie_data["active_time"] = active_time
                    with open(cookie_file, "w") as f:
                        json.dump(cookie_data, f)

                    print(f"\033[1;92m✅ Account {uid} active for {active_time}\033[0m")
                else:
                    print("\033[1;91m⚠️ Session Expired.\033[0m")
                    return

                time.sleep(interval)

            except requests.RequestException as e:
                print(f"\033[1;91m⚠️ Error during keep-alive: {e}\033[0m")
                max_retries -= 1
                if max_retries <= 0:
                    print("\033[1;91m✖️ Max retries exceeded. Exiting.\033[0m")
                    return
                time.sleep(5)

def main():
    os.system("clear")
    handler = FacebookAccountHandler()
    account_email = input("\033[1;92mEnter email for account: \033[0m").strip()
    custom_password = input("\033[1;93mType Your Password: \033[0m").strip()

    if custom_password:
        password = custom_password
    else:
        password = None  # Will trigger the BASE_PASSWORD + random suffix in generate_user_details

    result = handler.create_account(account_email, password=password)

    if result:
        uid, profile_link, account_password = result
        handler.keep_alive(uid, account_password)


if __name__ == "__main__":
    main()
